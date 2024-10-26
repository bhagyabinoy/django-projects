import json
import openpyxl
from django.http import HttpResponse
from food.models import FoodItem
from openpyxl.styles import Alignment, Font
from rest_framework import status
from rest_framework.permissions import AllowAny, IsAuthenticated
from rest_framework.response import Response
from rest_framework.views import APIView

from .models import *
from .serializers import *
from food.serializers import *
from django.shortcuts import render
from django.http import JsonResponse

import pdfkit
from django.http import HttpResponse
from django.template.loader import render_to_string
from .mixins import FilterMixins
from .utils import accounts_utils,logs_utils
from food.utils import food_utils, log_utils

##################################################################################################################################


class AddConsumed(APIView):

    permission_classes = [AllowAny]

    def post(self, request):         
        
        request_data = request.data.get("consumed_list")  
        if (type(request_data) == str):
            request_data = json.loads(request_data)
        if request_data is None:
            print("none")
            return Response("no input received", status=status.HTTP_400_BAD_REQUEST)

        else: 
            consumed_list = []   
            for data in request_data:           #iterating through request_data
                food_name = (data['food'])          #fetching name of food(s) from request_data 
                #food_queryset = FoodItem.objects.filter(name=food_name).values()            #accessing object wrt food name    
                food_queryset = accounts_utils.get_fooddetails_by_name(food_name)
                if food_queryset.count() == 0:		
                        return Response("This Food item is not in the list", status=status.HTTP_404_NOT_FOUND)     
              
                else: 
                    for item in food_queryset:
                        quantity = item['quantity']            #accessing calories and quantity from object
                        calories = item['calories']
                        calories_by_quantity = calories/quantity            #dividing calories by quantity to get calorie for 1 gm
                        data["calories"]= calories_by_quantity * data['quantity']             #multiplying above with quantity consumed by user & adding it to request_data dict
                        consumed_list.append(data)          #appending calories_consumed to list
                print(consumed_list, "consumed_list")
            serializer =  AddCaloriesConsumedSerializer(data=consumed_list, many=True)
            if serializer.is_valid(raise_exception=ValueError):
            
                serializer.save()
                return Response(serializer.data,  status=status.HTTP_201_CREATED)
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)


##################################################################################################################################


class ConsumeList(APIView):

    permission_classes = [AllowAny]

    def get(self, request): 
            
        #queryset = ConsumedByUser.objects.all().order_by('id')
        queryset = accounts_utils.get_all_consumedbyuser_instances()
        serializer = CaloriesConsumedSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)
    

##################################################################################################################################


class ConsumeDelete(APIView):

    permission_classes = ([IsAuthenticated])
  
    def get(self, request, pk):
        queryset = accounts_utils.consumed_detail(pk)
        queryset.delete()
        return Response("Data deleted successfully", status=status.HTTP_200_OK)


##################################################################################################################################


class SearchConsumedByDay(APIView, FilterMixins):
    permission_classes = [AllowAny]

    def get(self, request): 
        date = self.request.query_params.get('date')
        user = self.request.query_params.get('user')

        if date is None and  user is None:
            return Response("no input received", status=status.HTTP_400_BAD_REQUEST)
        else:
            queryset = ConsumedByUser.objects.filter(user=user, date=date)
            print("queryset",queryset)
            if queryset.count() == 0:		
                return Response("no records found", status=status.HTTP_404_NOT_FOUND)
            
            else:
                serializer = CaloriesConsumedSerializer(queryset, many=True)
                consumed_list=serializer.data
                print(queryset)

                calories_list = list(ConsumedByUser.objects.values_list('calories', flat=True).filter(user=user, date=date))
                print(calories_list)
                print(type(calories_list))
                net_calories = sum(calories_list)
                rounded_net_calories = round(net_calories)
                print(rounded_net_calories)
                context = {
                'net_calories': rounded_net_calories,
                'consumed_list': consumed_list,
                }
                return Response(context,  status=status.HTTP_200_OK)



##################################################################################################################################


class SearchConsumedByMonth(APIView, FilterMixins):

    permission_classes = [AllowAny]

    def get(self, request): 
        month  = request.query_params.get('month')
        user = request.query_params.get('user')
        year = request.query_params.get('year')

        if month is None: 
            return Response("Please enter month", status=status.HTTP_400_BAD_REQUEST)
        elif year is None: 
            return Response("Enter the year", status=status.HTTP_400_BAD_REQUEST)
        else: 

            
            consumed_list = self.consumed_by_month_mixin(month, year, user)

            #calories_list = list(ConsumedByUser.objects.values_list('calories', flat=True).filter(user=user, date__month=month))
            calories_list = list(accounts_utils.consumed_by_month(user,month))
            net_calories = sum(calories_list)
            # consumed_list = list(consumed_list) or []
            # net_calories = sum(item['calories'] for item in consumed_list)
            context = {
            'net_calories': net_calories,
            'consumed_list': consumed_list,
            }
            return Response(context,  status=status.HTTP_200_OK)           


##################################################################################################################################


class SearchConsumedByWeek(APIView, FilterMixins):

    permission_classes = [AllowAny]

    def get(self, request): 

        week  = request.query_params.get('week')
        user = request.query_params.get('user')
        year = request.query_params.get('year')

        if week is None: 
            return Response("Please enter a valid week number", status=status.HTTP_400_BAD_REQUEST)
        elif year is None: 
            return Response("Enter the year", status=status.HTTP_400_BAD_REQUEST)
        else:
            data = self.consumed_by_week_mixin(week, year, user)

            # calories_list = list(ConsumedByUser.objects.values_list('calories', flat=True).filter(user=user, date__week=week))
            calories_list = list(accounts_utils.consumed_by_week(user,week))
            net_calories = sum(calories_list)
            context = {
            'net_calories': net_calories,
            'consumed_list': data,
            }
            return Response(context,  status=status.HTTP_200_OK)


##################################################################################################################################


class SearchRange(APIView):

    permission_classes = [AllowAny]

    def get(self, request, pk): 
        
        start_date = request.data.get("start_date")
        end_date = request.data.get("end_date")
        user = request.data.get("user")
        queryset = ConsumedByUser.objects.all().filter(user=user, date__range=(start_date, end_date))
        serializer = CaloriesConsumedSerializer(queryset, many=True)
        return Response(serializer.data, status=status.HTTP_200_OK)


##################################################################################################################################

"""day wise exporting"""
class Export(APIView):

    permission_classes = [AllowAny]

    def get(self, request):

        date = request.query_params.get('date')
        user = request.query_params.get('user')
        print(date)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="calorie.xls"'

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Calorie Data"
        sheet.merge_cells('A1:E1')
        top_left_cell = sheet['A1']
        date_cell = sheet['D2']
        date_cell.value = "date"
        date_cell = sheet['E2']
        date_cell.value = date
        top_left_cell.font  = Font(b=True, color="000000FF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell.value = "Quantity and Calories Consumed"
        sheet.column_dimensions["A"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions["B"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions["C"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        
        headings = ['Food Item', 'Quantity Consumed', 'Calories']
        sheet.append(headings)

        #consumed_list = list(ConsumedByUser.objects.all().filter(user=user, date=date).values_list('food','quantity','calories'))
        consumed_list = list(accounts_utils.consumedlist_by_date(user,date))
        food_id_list = [x[0] for x in consumed_list]
        calories_list = [x[2] for x in consumed_list]
        quantity_list = [x[1] for x in consumed_list]
        food_name_list=[]
        for i in food_id_list:
            food_name_list.append(list(FoodItem.objects.values_list('name', flat=True).filter(id=i)))
        flat_list = [item for sublist in food_name_list for item in sublist]
        rows = list(zip(flat_list, calories_list, quantity_list))
        net_calories = sum(calories_list)
        print(net_calories)

        for row in rows:
            sheet.append(row)
        sheet.cell(row=len(consumed_list)+4, column=len(headings)-1).value = "Total Calories" 
        sheet.cell(row=len(consumed_list)+4, column=len(headings)   ).value = net_calories

        wb.save(response)
        return response


##################################################################################################################################

"""month wise exporting"""
class ExportMonth(APIView):

    permission_classes = [AllowAny]

    def get(self, request):

        month  = request.query_params.get('month')
        user = request.query_params.get('user')
        year = request.query_params.get('year')
        print("month:   ",month )
        print("user: ",user)
        print("year:   ",year)
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="caloriemonth.xls"'

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Calorie Data"
        sheet.merge_cells('A1:G1')
        top_left_cell = sheet['A1']
        date_cell = sheet['E2']
        date_cell.value = "Year"
        date_cell = sheet['F2']
        date_cell.value = year
        date_cell = sheet['G2']
        date_cell.value = "Month:"
        date_cell = sheet['H2']
        date_cell.value = month
        top_left_cell.font  = Font(b=True, color="000000FF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell.value = "Quantity and Calories Consumed"
        sheet.column_dimensions["A"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions["B"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions["C"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions["D"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        
        headings = ['Food Item', 'Quantity Consumed', 'Calories', 'Date']
        sheet.append(headings)
        #consumed_list = list(ConsumedByUser.objects.all().filter(user=user, date__year=year,date__month=month).values_list('food','quantity','calories', 'date'))
        consumed_list = list(accounts_utils.consumedlist_by_month(user,year, month))
        food_id_list = [x[0] for x in consumed_list]
        calories_list = [x[2] for x in consumed_list]
        quantity_list = [x[1] for x in consumed_list]
        date_list = [x[3] for x in consumed_list]
        date_list=[date.strftime('%Y-%m-%d') for date in date_list]
        food_name_list=[]
        for i in food_id_list:
            food_name_list.append(list(FoodItem.objects.values_list('name', flat=True).filter(id=i)))
        flat_list = [item for sublist in food_name_list for item in sublist]
        rows = list(zip(flat_list, calories_list, quantity_list, date_list))
        net_calories = sum(calories_list)

        for row in rows:
            sheet.append(row)
        sheet.cell(row=len(consumed_list)+4, column=len(headings)+1).value = "Total Calories" 
        sheet.cell(row=len(consumed_list)+4, column=len(headings)+2  ).value = net_calories

        wb.save(response)
        return response
  


##################################################################################################################################

class ExportWeek(APIView):


    permission_classes = [AllowAny]

    def get(self, request):

        week  = request.query_params.get('week')
        user = request.query_params.get('user')
        year = request.query_params.get('year')
        response = HttpResponse(content_type='application/ms-excel')
        response['Content-Disposition'] = 'attachment; filename="calorieweek.xls"'

        wb = openpyxl.Workbook()
        sheet = wb.active
        sheet.title = "Calorie Data"
        sheet.merge_cells('A1:G1')
        top_left_cell = sheet['A1']
        date_cell = sheet['E2']
        date_cell.value = "Year"
        date_cell = sheet['F2']
        date_cell.value = year
        date_cell = sheet['G2']
        date_cell.value = "Weak:"
        date_cell = sheet['H2']
        date_cell.value = week
        top_left_cell.font  = Font(b=True, color="000000FF")
        top_left_cell.alignment = Alignment(horizontal="center", vertical="center")
        top_left_cell.value = "Quantity and Calories Consumed"
        sheet.column_dimensions["A"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions["B"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions["C"].alignment = Alignment(horizontal='center')
        sheet.column_dimensions['A'].width = 20
        sheet.column_dimensions['B'].width = 20
        sheet.column_dimensions['C'].width = 20
        sheet.column_dimensions['D'].width = 20
        sheet.column_dimensions['E'].width = 20
        
        headings = ['Food Item', 'Quantity Consumed', 'Calories', 'Date']
        sheet.append(headings)

        
        #consumed_list = list(ConsumedByUser.objects.all().filter(user=user, date__year=year,date__week=week).values_list('food','quantity','calories', 'date'))
        consumed_list = list(accounts_utils.consumedlist_by_week(user,year, week))
        food_id_list = [x[0] for x in consumed_list]
        calories_list = [x[2] for x in consumed_list]
        quantity_list = [x[1] for x in consumed_list]
        date_list = [x[3] for x in consumed_list]
        date_list=[date.strftime('%Y-%m-%d') for date in date_list]
        food_name_list=[]
        for i in food_id_list:
            food_name_list.append(list(FoodItem.objects.values_list('name', flat=True).filter(id=i)))
        flat_list = [item for sublist in food_name_list for item in sublist]
        rows = list(zip(flat_list, calories_list, quantity_list, date_list))
        net_calories = sum(calories_list)
        print(net_calories)

        for row in rows:
            sheet.append(row)
        sheet.cell(row=len(consumed_list)+4, column=len(headings)+1).value = "Total Calories" 
        sheet.cell(row=len(consumed_list)+4, column=len(headings)+2  ).value = net_calories

        wb.save(response)
        return response
  


##################################################################################################################################

class Convert_pdf_by_date(APIView):


    def get(self, request): 
        date = request.query_params.get('date')
        user = request.query_params.get('user')
        if date is None and  user is None:
            return Response("no input received", status=status.HTTP_400_BAD_REQUEST)
        else:
            #consumed_list = list(ConsumedByUser.objects.all().filter(user=user, date=date).values_list('food','quantity','calories', 'date'))
            consumed_list = list(accounts_utils.consumedlist_by_date(user,date))   
            food_id_list = [x[0] for x in consumed_list]
            calories_list = [x[2] for x in consumed_list]
            quantity_list = [x[1] for x in consumed_list]
            date_list = [x[3] for x in consumed_list]
            food_name_list=[]
            for i in food_id_list:
                food_name_list.append(list(FoodItem.objects.values_list('name', flat=True).filter(id=i)))
            flat_list = [item for sublist in food_name_list for item in sublist]
            foodlist = list(zip(flat_list, calories_list, quantity_list, date_list))
            net_calories = sum(calories_list)
            table_data={}
            table_data['foodlist'] = foodlist
            table_data['net_calories'] = net_calories
            html_string = render_to_string('pdfsheet.html', table_data)
            result = pdfkit.from_string(html_string)
            response = HttpResponse(result, content_type='application/pdf')
            response['Content-Disposition']='attachment;filename="foodlist.pdf"'
            return response
            

##################################################################################################################################

class Convert_pdf_by_week(APIView):


    def get(self, request): 
        week  = request.query_params.get('week')
        user = request.query_params.get('user')
        year = request.query_params.get('year')        
        if week is None: 
            return Response("Please enter a valid week number", status=status.HTTP_400_BAD_REQUEST)
        elif year is None: 
            return Response("Enter the year", status=status.HTTP_400_BAD_REQUEST)
        elif week.isnumeric() and year.isnumeric():
            
            #consumed_list = list(ConsumedByUser.objects.all().filter(user=user, date__year=year,date__week=week).values_list('food','quantity','calories', 'date'))
            consumed_list = list(accounts_utils.consumedlist_by_week(user,year, week))
            food_id_list = [x[0] for x in consumed_list]
            calories_list = [x[2] for x in consumed_list]
            quantity_list = [x[1] for x in consumed_list]
            date_list = [x[3] for x in consumed_list]
            print(date_list)
            food_name_list=[]
            for i in food_id_list:
                food_name_list.append(list(FoodItem.objects.values_list('name', flat=True).filter(id=i)))
            flat_list = [item for sublist in food_name_list for item in sublist]
            foodlist = list(zip(flat_list, calories_list, quantity_list, date_list))
            net_calories = sum(calories_list)
            table_data={}
            table_data['foodlist'] = foodlist
            table_data['net_calories'] = net_calories
            html_string = render_to_string('pdfsheet.html', table_data)
            result = pdfkit.from_string(html_string)
            response = HttpResponse(result, content_type='application/pdf')
            response['Content-Disposition']='attachment;filename="foodlist.pdf"'
            return response
        else:
            return Response("Please enter a valid integer", status=status.HTTP_400_BAD_REQUEST)    

##################################################################################################################################


class Convert_pdf_by_month(APIView):


    def get(self, request): 
        month  = request.query_params.get('month')
        user = request.query_params.get('user')
        year = request.query_params.get('year')
        if month is None: 
            return Response("Please enter month", status=status.HTTP_400_BAD_REQUEST)
        elif year is None: 
            return Response("Enter the year", status=status.HTTP_400_BAD_REQUEST)
        elif month.isnumeric() and year.isnumeric():
            #consumed_list = list(ConsumedByUser.objects.all().filter(user=user, date__year=year,date__month=month).values_list('food','quantity','calories', 'date'))
            consumed_list = list(accounts_utils.consumedlist_by_month(user,year, month))
            food_id_list = [x[0] for x in consumed_list]
            calories_list = [x[2] for x in consumed_list]
            quantity_list = [x[1] for x in consumed_list]
            date_list = [x[3] for x in consumed_list]
            date_list=[date.strftime('%Y-%m-%d') for date in date_list]
            food_name_list=[]
            for i in food_id_list:
                food_name_list.append(list(FoodItem.objects.values_list('name', flat=True).filter(id=i)))
            flat_list = [item for sublist in food_name_list for item in sublist]
            foodlist = list(zip(flat_list, calories_list, quantity_list, date_list))
            net_calories = sum(calories_list)
            table_data={}
            table_data['foodlist'] = foodlist
            table_data['net_calories'] = net_calories
            html_string = render_to_string('pdfsheet.html', table_data)
            result = pdfkit.from_string(html_string)
            response = HttpResponse(result, content_type='application/pdf')
            response['Content-Disposition']='attachment;filename="foodlist.pdf"'
            return response
        else:
            return Response("Please enter a valid integer", status=status.HTTP_400_BAD_REQUEST)   


    
##################################################################################################################################

class weekwisechart_mixin(APIView, FilterMixins):

    def get(self, request): 

        week  = request.query_params.get('week')
        user = request.query_params.get('user')
        year = request.query_params.get('year')
        print(week, user, year)

        data = self.consumed_by_week_mixin(week, year, user)

        calories = [d['calories'] for d in data]
        quantity = [d['quantity'] for d in data]
        food_name = [d['food'] for d in data] 
        chartLabel = "Calories"
        veg=[]
        nonveg=[]
        total_veg=0
        total_nonveg=0
        for name in food_name:
            #queryset_food = FoodItem.objects.get(name=name)
            queryset_food = accounts_utils.get_food_details(name)
            serializer = FoodSerializer(queryset_food, many=False)
            food_detail=serializer.data
            print(food_detail)
            food_quantity= food_detail.get('quantity')
            if(food_detail.get('category')=='Veg'):
                print("Veg")
                veg.append(food_quantity)
            else:
                nonveg.append(food_quantity)
        print(veg)
        print(nonveg)
        for ele in range(0, len(veg)):
            total_veg = total_veg + veg[ele]
        for ele in range(0, len(nonveg)):
            total_nonveg = total_nonveg + nonveg[ele]
        print(total_veg)
        print(total_nonveg)
        labels_piechart =  ["Veg", "Non Veg"]  
        chartLabel_piechart = "Quantity Consumed based on category"
        chartdata_piechart = [total_veg, total_nonveg] 
        print(labels_piechart)
        print(chartLabel_piechart)
        print(chartdata_piechart)

        data ={
                    "food_name":food_name,
                    "chartLabel":chartLabel,
                    "calories":calories,
                    "quantity":quantity,
                    "labels_piechart": labels_piechart,
                    "chartLabel_piechart":chartLabel_piechart,
                    "chartdata_piechart":chartdata_piechart
            }
        print(data)
        return JsonResponse(data)

##################################################################################################################################

class daywisechart_mixin(APIView, FilterMixins):

    permission_classes = [AllowAny]

    def get(self, request): 

        date = self.request.query_params.get('date')
        user = self.request.query_params.get('user')

    
        data = self.consumed_by_date_mixin(date, user)

        calories = [d['calories'] for d in data]
        quantity = [d['quantity'] for d in data]
        food_name = [d['food'] for d in data] 
        chartLabel = "Calories"
        veg=[]
        nonveg=[]
        total_veg=0
        total_nonveg=0
        for name in food_name:
            #queryset_food = FoodItem.objects.get(name=name)
            queryset_food = accounts_utils.get_food_details(name)
            #print(queryset_food)
            # print(type(queryset_food))
            serializer = FoodSerializer(queryset_food, many=False)
            food_detail=serializer.data
            print(food_detail)
            food_quantity= food_detail.get('quantity')
            if(food_detail.get('category')=='Veg'):
                print("Veg")
                veg.append(food_quantity)
            else:
                nonveg.append(food_quantity)
                print(veg)
                print(nonveg)
                for ele in range(0, len(veg)):
                    total_veg = total_veg + veg[ele]
                for ele in range(0, len(nonveg)):
                    total_nonveg = total_nonveg + nonveg[ele]
                print(total_veg)
                print(total_nonveg)
        labels_piechart =  ["Veg", "Non Veg"]  
        chartLabel_piechart = "Quantity Consumed based on category"
        chartdata_piechart = [total_veg, total_nonveg] 
        print(labels_piechart)
        print(chartLabel_piechart)
        print(chartdata_piechart)

        data ={
                    "food_name":food_name,
                    "chartLabel":chartLabel,
                    "calories":calories,
                    "quantity":quantity,
                    "labels_piechart": labels_piechart,
                    "chartLabel_piechart":chartLabel_piechart,
                    "chartdata_piechart":chartdata_piechart
            }
        print(data)
        return JsonResponse(data)
    

##################################################################################################################################


class monthwisechart_mixin(APIView, FilterMixins):

    permission_classes = [AllowAny]

    def get(self, request): 

        month  = request.query_params.get('month')
        user = request.query_params.get('user')
        year = request.query_params.get('year')

        data = self.consumed_by_month_mixin((month, year, user))

        calories = [d['calories'] for d in data]
        quantity = [d['quantity'] for d in data]
        food_name = [d['food'] for d in data] 
        chartLabel = "Calories"
        veg=[]
        nonveg=[]
        total_veg=0
        total_nonveg=0
        for name in food_name:
            #queryset_food = FoodItem.objects.get(name=name)
            queryset_food = accounts_utils.get_food_details(name)   
            #print(queryset_food)
            # print(type(queryset_food))
            serializer = FoodSerializer(queryset_food, many=False)
            food_detail=serializer.data
            print(food_detail)
            food_quantity= food_detail.get('quantity')
            if(food_detail.get('category')=='Veg'):
                print("Veg")
                veg.append(food_quantity)
            else:
                nonveg.append(food_quantity)
                print(veg)
                print(nonveg)
                for ele in range(0, len(veg)):
                    total_veg = total_veg + veg[ele]
                for ele in range(0, len(nonveg)):
                    total_nonveg = total_nonveg + nonveg[ele]
                print(total_veg)
                print(total_nonveg)
        labels_piechart =  ["Veg", "Non Veg"]  
        chartLabel_piechart = "Quantity Consumed based on category"
        chartdata_piechart = [total_veg, total_nonveg] 
        print(labels_piechart)
        print(chartLabel_piechart)
        print(chartdata_piechart)

        data ={
                    "food_name":food_name,
                    "chartLabel":chartLabel,
                    "calories":calories,
                    "quantity":quantity,
                    "labels_piechart": labels_piechart,
                    "chartLabel_piechart":chartLabel_piechart,
                    "chartdata_piechart":chartdata_piechart
            }
        print(data)
        return JsonResponse(data)
    
##################################################################################################################################


def daywisechartpage(request):
    return render(request, 'daywise_chart.html')

def weekwisechartpage(request):
    return render(request, 'weekwise_chart.html')

def monthwisechartpage(request):
    return render(request, 'monthwise_chart.html')


    
##################################################################################################################################

